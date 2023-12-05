import pyrealsense2 as rs
import numpy as np
import cv2
import matlab.engine
import math
import featureextraction_ShiTomas
import openpyxl
import time

# adding new parts to the database
# 1 make a screenshot to the part in the instructions manual
# 2 generating edge map of it
# 3 add the points information to the dictionary
# 4 add size information to the dictionary
# 5 add color information to the dictionary

coordinates = np.zeros((4, 2), np.int32)

pipeline = rs.pipeline()
config = rs.config()
config.enable_stream(rs.stream.depth, 640, 480, rs.format.z16, 30)
config.enable_stream(rs.stream.color, 640, 480, rs.format.bgr8, 30)
# Start streaming
pipeline.start(config)
align_to = rs.stream.color
align = rs.align(align_to)

index=1

def backgroundcolorchoose(image):
    sizea=len(image)
    sizeb=len(image[0])
    n=sizea/2
    m=sizeb/2
    grayforbackground = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    if grayforbackground[int(n)][int(m)] <110 :
        backcolor='black'
        return backcolor
    else:
        backcolor='white'
        return backcolor
def camera_pointselect():
    counter = 1
    while True:
        # Wait for a coherent pair of frames: depth and color
        frames = pipeline.wait_for_frames()

        depth_frame = frames.get_depth_frame()
        color_frame = frames.get_color_frame()

        if not depth_frame or not color_frame:
            continue
        # Convert images to numpy arrays
        color_image = np.asanyarray(color_frame.get_data())
        cv2.imshow('RealSense', color_image)
        key = cv2.waitKey(1)
        if key == ord('q'):
            if counter == 1:
                backgroundcolor=backgroundcolorchoose(color_image)
                print(backgroundcolor)
                counter=counter+1
            else:
                coordinates=featureextraction_ShiTomas.main(backgroundcolor,color_image)
                cv2.destroyAllWindows()
                break

    return coordinates,color_image # (x,y)

def get_aligned_images():
    frames = pipeline.wait_for_frames()  # 等待获取图像帧，获取颜色和深度的框架集
    aligned_frames = align.process(frames)  # 获取对齐帧，将深度框与颜色框对齐

    aligned_depth_frame = aligned_frames.get_depth_frame()  # 获取对齐帧中的的depth帧
    aligned_color_frame = aligned_frames.get_color_frame()  # 获取对齐帧中的的color帧

    #### 获取相机参数 ####
    depth_intrin = aligned_depth_frame.profile.as_video_stream_profile().intrinsics  # 获取深度参数（像素坐标系转相机坐标系会用到）
    color_intrin = aligned_color_frame.profile.as_video_stream_profile().intrinsics  # 获取相机内参

    #### 将images转为numpy arrays ####
    img_color = np.asanyarray(aligned_color_frame.get_data())  # RGB图
    img_depth = np.asanyarray(aligned_depth_frame.get_data())  # 深度图（默认16位）

    return color_intrin, depth_intrin, img_color, img_depth, aligned_depth_frame

def get_3d_camera_coordinate(depth_pixel, aligned_depth_frame, depth_intrin):
    x = depth_pixel[0]
    y = depth_pixel[1]
    dis = aligned_depth_frame.get_distance(x, y)  # 获取该像素点对应的深度
    # print ('depth: ',dis)       # 深度单位是m
    camera_coordinate = rs.rs2_deproject_pixel_to_point(depth_intrin, depth_pixel, dis)
    # print ('camera_coordinate: ',camera_coordinate)
    return dis, camera_coordinate

if __name__ == "__main__":
    # --------------------------------
    # Recognition part
    # --------------------------------
    #query part for homography

    # Number of parts in this group

    n = 8
    test_time=50
    pointseleceted_dictionary = np.load('pointselected_twochair_dictionary_smallsize.npy', allow_pickle=True).item()
    objsize = np.load('objsize.npy', allow_pickle=True).item()
    eng = matlab.engine.start_matlab()
    for ex_times in range(test_time):
        starttime = time.time()
        print(ex_times)
        if ex_times == 0:
            coordinates,image1=camera_pointselect()
        gray = cv2.cvtColor(image1, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, 50, 150, apertureSize=3)
        pts_src = np.float32([coordinates[0], coordinates[1], coordinates[2], coordinates[3]])
        im_src = edges.copy()
        # homography transformation and FDCM in matlab
        partnumber_matched=''
        shapecost=np.zeros(n,dtype=float)
        times = 10       #times of the img which used in define the difference between the stantard bounding box and recognition result
            # --------------------------------
            #attribute: shape(contour) by using fdcm_previous
            # --------------------------------

        for i in range(n):
            j=i+1

            size_max = np.amax(pointseleceted_dictionary[j], axis=0)
            size_min = np.amin(pointseleceted_dictionary[j], axis=0)
            size_max_x = int(size_max[0])
            size_max_y = int(size_max[1])
            size_min_x = int(size_min[0])
            size_min_y = int(size_min[1])

            #homography transformation
            pts_dst = np.float32(pointseleceted_dictionary[j])
            im_dst=cv2.imread("C:/Users/wangq/PycharmProjects/pythonProject/intelrealsense_wcx/templateimage/smallsize/" + str(j) + ".png")
            h, status = cv2.findHomography(pts_src, pts_dst)
            im_out = cv2.warpPerspective(im_src, h, (im_dst.shape[1], im_dst.shape[0]))

            sav = str(j)+'.pgm'
            cv2.imwrite(sav, im_out)
            # matlab :fast directional chamfer matching
            partnumber ,value = eng.chamfermatching_nopic_forpython(sav, j, size_min_x, size_min_y, size_max_x, size_max_y,times , nargout=2)
            print("partnumber: ",j, "shapecost: ",value)

            shapecost[i]=value

            # --------------------------------
            # attribute: size
            # --------------------------------
        camera_coordinates = np.zeros((4, 3), dtype=float)
        camera_coordinate_matrix  = np.zeros(3, dtype = float)
        camera_coordinate = np.zeros(3, dtype = float)
        boxsize=4 # halfsize
        # the box of computing the average points
        for i in range(4):
            # Get the matched depth img and color img
            color_intrin, depth_intrin, img_color, img_depth, aligned_depth_frame = get_aligned_images()  # 获取对齐图像与相机参数
            depth_pixel = coordinates[i]
        # compute 9 points' average position
            count = 0
            camera_coordinate = np.zeros(3, dtype = float)
            distance=0
            for j in range(-boxsize, 1 + boxsize):
                for k in range(-boxsize, 1 + boxsize):
                    dis, camera_coordinate_matrix = get_3d_camera_coordinate([int(depth_pixel[0]) + j, int(depth_pixel[1]) + k],
                                                                             aligned_depth_frame, depth_intrin)
                    if camera_coordinate_matrix[2] != 0:
                        count += 1
                        camera_coordinate = camera_coordinate + camera_coordinate_matrix
                    distance=dis+distance
            camera_coordinate = camera_coordinate / count
            camera_coordinates[i][:] = camera_coordinate
        obj_legth = math.sqrt((camera_coordinates[0][0] - camera_coordinates[1][0]) ** 2 + (
                    camera_coordinates[0][1] - camera_coordinates[1][1]) ** 2 + (
                                          camera_coordinates[0][2] - camera_coordinates[1][2]) ** 2)
        obj_width = math.sqrt((camera_coordinates[1][0] - camera_coordinates[2][0]) ** 2 + (
                    camera_coordinates[1][1] - camera_coordinates[2][1]) ** 2 + (
                                          camera_coordinates[1][2] - camera_coordinates[2][2]) ** 2)

        print("bounding box limit: ",times, "boxsize: ", boxsize, "Distance: ", dis)
        print("object length is: %.4f"% obj_legth,"  object width is: %.4f"% obj_width)

        #params in final cost compotation
        A=1
        B=1
        chosen=100
        chosencost = 100
        sizecost=np.zeros(n,dtype=float)
        for i in range(n):
            j=i+1
            size=objsize[j]
            sizecost[i]=abs(size[0]-obj_legth)+abs(size[1]-obj_width)
            if(np.isnan(sizecost[i])):
                sizecost[i]=0.5
            cost=A*sizecost[i]+B*shapecost[i]
            if(chosencost>cost):
                chosen=j
                chosencost=cost

            print("partnumber: %d"% j,"  sizecost: %.4f"% sizecost[i], "  shapecost: %.4f"% shapecost[i],"  cost: %.4f"% cost)
        print("The recognition result is(part number): ",chosen," cost is : %.4f"% chosencost)
        endtime = time.time()
        lasttime=endtime-starttime
        # data saving to excel

        excel_name = 'exdata.xlsx'
        wb = openpyxl.load_workbook(excel_name)
        ws = wb['07210601']
        m = 1
        # n = ws.max_row
        ws.cell(row=ex_times+1+m, column=1, value=str(ex_times+1))
        ws.cell(row=ex_times+1+m, column=2, value=str(chosen))
        ws.cell(row=ex_times + 1 + m, column=3, value=str(chosencost))
        ws.cell(row=ex_times + 1 + m, column=4, value=str(shapecost[chosen-1]))
        ws.cell(row=ex_times + 1 + m, column=5, value=str(sizecost[chosen-1]))
        ws.cell(row=ex_times + 1 + m, column=6, value=str((shapecost[0]+sizecost[0])))
        ws.cell(row=ex_times + 1 + m, column=7, value=str((shapecost[1]+sizecost[1])))
        ws.cell(row=ex_times + 1 + m, column=8, value=str((shapecost[2]+sizecost[2])))
        ws.cell(row=ex_times + 1 + m, column=9, value=str((shapecost[3]+sizecost[3])))
        ws.cell(row=ex_times + 1 + m, column=10, value=str((shapecost[4]+sizecost[4])))
        ws.cell(row=ex_times + 1 + m, column=11, value=str((shapecost[5]+sizecost[5])))
        ws.cell(row=ex_times + 1 + m, column=12, value=str((shapecost[6] + sizecost[6])))
        ws.cell(row=ex_times + 1 + m, column=12, value=str((shapecost[7] + sizecost[7])))
        ws.cell(row=ex_times + 1 + m, column=13, value=str(lasttime))


        wb.save('exdata.xlsx')
        # --------------------------------
        # Position detection part
        # --------------------------------