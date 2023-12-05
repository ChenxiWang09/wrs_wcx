
import openpyxl

excel_name = 'exdata.xlsx'
wb = openpyxl.load_workbook(excel_name)
ws=wb['07060301']
n=ws.max_column
m=ws.max_row
ws.cell(row=1, column=1, value=str(1))
wb.save('exdata.xlsx')
